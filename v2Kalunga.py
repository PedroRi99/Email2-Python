from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from email.message import EmailMessage
from tkinter import messagebox
import tkinter as tk
import ssl
import smtplib
import openpyxl
import time
import pyperclip

nav = webdriver.Chrome()
nav.get("https://www.kalunga.com.br/busca/1?q=teclado")

# //tag[@atributo='valor']

titulos = nav.find_elements(By.XPATH, "//h2[@class='blocoproduto__title mb-0 mt-2 pb-2 pb-lg-3']")
precos = nav.find_elements(By.XPATH, "//span[@class='blocoproduto__text blocoproduto__text--bold blocoproduto__price']")
link = nav.find_elements(By.XPATH, "//a[@class='blocoproduto__link']")

workbook = openpyxl.Workbook()
workbook.create_sheet('produtos')
sheetProdutos = workbook['produtos']
sheetProdutos['A1'].value = 'Nome'
sheetProdutos['B1'].value = 'Preço'
sheetProdutos['C1'].value = 'Links'

for titulo, preco, site in zip(titulos, precos, link):
    sheetProdutos.append([titulo.text, preco.text, site.get_attribute('href')])
workbook.save('produtos.xlsx')

def enviarEmail(body):
    planilha = openpyxl.load_workbook('produtos.xlsx')
    sheet = planilha['produtos']

    meuEmail = ""
    senha = ""

    assunto = "Teste envio e-mail - python"
    em = EmailMessage()

    em['From'] = meuEmail
    em['To'] = sheet['D2'].value
    em['subject'] = assunto
    em.set_content(body)

    context = ssl.create_default_context()

    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(meuEmail, senha)
        smtp.sendmail(meuEmail, sheet['D2'].value, em.as_string())

#TK

def usuarioTK():
    janela = tk.Tk()
    janela.title("Email Usuário")
    janela.geometry("400x400")

    label_usuario = tk.Label(janela, text="Usuário:")
    label_usuario.pack()

    entrada_usuario = tk.Entry(janela, width=50)
    entrada_usuario.pack()

    def returnUser():
        planilha = openpyxl.load_workbook('produtos.xlsx')
        sheet = planilha['produtos']
        cell = sheet['D2']
        cell.value = entrada_usuario.get()
        planilha.save('produtos.xlsx')

        janela.destroy()
        return
    
    botao_login = tk.Button(janela, text="Enviar", command=returnUser)
    botao_login.pack()

    janela.mainloop()

#EMAIL


def infoExcel():
    planilha = openpyxl.load_workbook('produtos.xlsx')
    sheet = planilha['produtos']

    for row in sheet.iter_rows(min_row=2):
        nome = row[0]
        preco = row[1]
        link = row[2]
        body = f"""
Ótimas promoções na kalunga de teclados!!
Nome do teclado: {nome.value}
Preço: {preco.value}
Link: {link.value}
"""
        
        planilha2 = openpyxl.load_workbook('produtos.xlsx')
        sheet2 = planilha2['produtos']
        if sheet2['D2'].value == None:
            usuarioTK()

        time.sleep(5)
        enviarEmail(body)

    
infoExcel()